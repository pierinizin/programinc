"""
Reestrutura o export do Kartado num boletim legível.

Duas abas:
  Boletim — um bloco por apontamento, para ler e imprimir
  Dados   — uma linha por SERVIÇO, para filtrar e girar em tabela dinâmica

O parser é dirigido pelo cabeçalho e aceita os dois formatos que o Kartado
exporta, sem alteração:
  · blocos repetidos  "Serviços (Manual) 1: Km Inicial", "... 2: ..."  (pintura)
  · colunas planas    m1..m10 + Média de Leituras                      (ensaios)

Uso:  python boletim.py entrada.xlsx saida.xlsx
"""
import sys
import re
from datetime import datetime

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Paleta do Incovia levada para o Excel.
TINTA  = 'FF1B2027'
FAIXA  = 'FFFFC72C'
CINZA  = 'FF6B7784'
LINHA  = 'FFD7DCE1'
CLARO  = 'FFF6F7F8'
BRANCO = 'FFFFFFFF'
ALERTA = 'FFB3401A'

FONTE = 'Arial'

IDENT = [
    'Serial', 'Serial Inventário Vinculado', 'Natureza', 'Status', 'Criado por',
    'Empresa', 'Equipe', 'Criado em', 'Encontrado em', 'Atualizado em',
    'Executado em', 'Cliente', 'Contrato', 'Encarregado', 'Local de Obra',
    'Fechamento de pista', 'Liberação de pista', 'Engenheiro responsável',
]

RE_BLOCO   = re.compile(r'^(.*?)\s(\d+):\s*(.+)$')
RE_LEITURA = re.compile(r'^m(\d+)$', re.I)
RE_FOTO    = re.compile(r'^foto\s*\d*$', re.I)

# Campo do bloco -> chave interna. Os nomes variam com a natureza
# ("Extensão de Pintura Manual" x "Extensão"), então o casamento é por prefixo.
CAMPOS = [
    (re.compile(r'^km inicial', re.I),            'kmIni'),
    (re.compile(r'^km final', re.I),              'kmFim'),
    (re.compile(r'^cor', re.I),                   'cor'),
    (re.compile(r'^local de apl', re.I),          'local'),
    (re.compile(r'^tipo de faixa', re.I),         'tipoFaixa'),
    (re.compile(r'^quantidade de faixas', re.I),  'faixas'),
    (re.compile(r'^extens', re.I),                'ext'),
    (re.compile(r'^largura', re.I),               'larg'),
    (re.compile(r'^[áa]rea total', re.I),         'area'),
    (re.compile(r'^observa', re.I),               'obs'),
    (re.compile(r'^tamanho da se', re.I),         'secao'),
    (re.compile(r'^cad', re.I),                   'cadencia'),
    (re.compile(r'^quantidade$', re.I),           'qtd'),
    (re.compile(r'^unidade', re.I),               'un'),
    (re.compile(r'^lado', re.I),                  'lado'),
]
NUMERICOS = {'kmIni', 'kmFim', 'faixas', 'ext', 'larg', 'area', 'secao', 'qtd'}


def chave_campo(nome):
    for regex, k in CAMPOS:
        if regex.match(nome.strip()):
            return k
    return None


def texto(v):
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%d/%m/%Y %H:%M')
    # o Kartado devolve vários campos com espaço sobrando ("EPR ", "Fabiano ")
    return str(v).strip()


def num(v):
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    s = texto(v)
    if not s:
        return None
    if s.count(','):
        s = s.replace('.', '').replace(',', '.')
    try:
        return float(s)
    except ValueError:
        return None


def data_br(v):
    return v.strftime('%d/%m/%Y') if isinstance(v, datetime) else texto(v)


def hora_br(v):
    return v.strftime('%H:%M') if isinstance(v, datetime) else texto(v)


def br(n, casas=2):
    return f'{n:,.{casas}f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


# ----------------------------------------------------------------------
# Leitura
# ----------------------------------------------------------------------
def ler(caminho):
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    aps = []

    for nome_aba in wb.sheetnames:
        ws = wb[nome_aba]
        linhas = [list(r) for r in ws.iter_rows(values_only=True)]
        if not linhas:
            continue

        i_head = next((i for i, l in enumerate(linhas)
                       if l and any(texto(c).lower() == 'serial' for c in l)), None)
        if i_head is None:
            continue
        head = [texto(h) for h in linhas[i_head]]

        # separa colunas simples das que pertencem a um bloco repetido
        simples, blocos = [], {}
        for i, h in enumerate(head):
            if not h:
                continue
            m = RE_BLOCO.match(h)
            if m:
                grupo, idx, campo = m.group(1), int(m.group(2)), m.group(3)
                blocos.setdefault(grupo, {}).setdefault(idx, []).append((i, campo))
            else:
                simples.append((i, h))

        grupo_serv = next((g for g in blocos if re.search(r'servi', g, re.I)),
                          next(iter(blocos), None))

        for linha in linhas[i_head + 1:]:
            if not linha:
                continue

            def pega(h, _linha=linha):
                return next((_linha[i] for i, x in simples if x == h), None)

            serial = texto(pega('Serial'))
            if not serial or serial.upper() == 'TOTAIS':
                continue

            ident = {h: pega(h) for h in IDENT}
            extras, leituras = {}, []
            for i, h in simples:
                if h in IDENT or RE_FOTO.match(h):
                    continue
                v = linha[i]
                if v is None or not texto(v):
                    continue
                m = RE_LEITURA.match(h)
                if m:
                    leituras.append((int(m.group(1)), v))
                else:
                    extras[h] = v

            servicos = []
            if grupo_serv:
                for idx in sorted(blocos[grupo_serv]):
                    s = {'n': idx}
                    for i, campo in blocos[grupo_serv][idx]:
                        v = linha[i]
                        if v is None or not texto(v):
                            continue
                        k = chave_campo(campo)
                        if k:
                            s[k] = num(v) if k in NUMERICOS else texto(v)
                        else:
                            s.setdefault('outros', {})[campo] = texto(v)
                    # bloco vazio é slot não usado do formulário, não serviço
                    if any(s.get(k) is not None for k in ('kmIni', 'ext', 'area', 'qtd')):
                        servicos.append(s)

            fotos = sum(1 for i, h in simples
                        if RE_FOTO.match(h) and linha[i] is not None and texto(linha[i]))

            aps.append({
                'aba': nome_aba, 'serial': serial, 'ident': ident, 'extras': extras,
                'servicos': servicos, 'leituras': sorted(leituras), 'fotos': fotos,
                'area': sum(s.get('area') or 0 for s in servicos),
                'ext': sum(s.get('ext') or 0 for s in servicos),
            })

    aps.sort(key=lambda a: (a['ident'].get('Executado em') or datetime.min, a['serial']))
    return aps


# ----------------------------------------------------------------------
# Estilos
# ----------------------------------------------------------------------
def borda(*lados):
    fina = Side(style='thin', color=LINHA)
    return Border(**{l: fina for l in lados})


def pinta(ws, r, cols, cor):
    for c in range(1, cols + 1):
        ws.cell(r, c).fill = PatternFill('solid', fgColor=cor)


def faixa_titulo(ws, r, esq, dir_, n):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n // 2)
    c = ws.cell(r, 1, esq)
    c.font = Font(name=FONTE, size=11, bold=True, color=BRANCO)
    c.alignment = Alignment(vertical='center', indent=1)
    ws.merge_cells(start_row=r, start_column=n // 2 + 1, end_row=r, end_column=n)
    c2 = ws.cell(r, n // 2 + 1, dir_)
    c2.font = Font(name=FONTE, size=9, color='FFD9DEE4')
    c2.alignment = Alignment(vertical='center', horizontal='right', indent=1)
    pinta(ws, r, n, TINTA)
    ws.row_dimensions[r].height = 20


def sub_faixa(ws, r, rotulo, n):
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=n)
    c = ws.cell(r, 1, rotulo.upper())
    c.font = Font(name=FONTE, size=8, bold=True, color=TINTA)
    c.alignment = Alignment(vertical='center', indent=1)
    pinta(ws, r, n, FAIXA)
    ws.row_dimensions[r].height = 15


def encurtar(rotulo):
    """O Kartado repete a natureza no nome do campo ("Tipo de Serviço - Lev. de
    Retrorrefletância"). A natureza já está no cabeçalho do bloco."""
    r = str(rotulo)
    if len(r) > 26 and ' - ' in r:
        r = r.split(' - ')[0]
    return r


def par(ws, r, rot_ini, rot_fim, val_ini, val_fim, rotulo, valor):
    ws.merge_cells(start_row=r, start_column=rot_ini, end_row=r, end_column=rot_fim)
    c = ws.cell(r, rot_ini, encurtar(rotulo).upper())
    c.font = Font(name=FONTE, size=7.5, bold=True, color=CINZA)
    c.alignment = Alignment(vertical='center', horizontal='right', indent=1)
    ws.merge_cells(start_row=r, start_column=val_ini, end_row=r, end_column=val_fim)
    v = ws.cell(r, val_ini, valor if texto(valor) else '—')
    v.font = Font(name=FONTE, size=9.5, color=TINTA)
    v.alignment = Alignment(vertical='center', indent=1)
    for col in range(rot_ini, val_fim + 1):
        ws.cell(r, col).border = borda('bottom')


# ----------------------------------------------------------------------
# Aba Boletim
# ----------------------------------------------------------------------
COLS_SERV = [
    ('#', 5), ('KM INI', 9), ('KM FIM', 9), ('COR', 11), ('LOCAL DE APLICAÇÃO', 19),
    ('TIPO DE FAIXA', 15), ('FAIXAS', 8), ('EXTENSÃO (m)', 12),
    ('LARGURA (m)', 11), ('ÁREA (m²)', 11), ('OBSERVAÇÕES', 28),
]
NC = len(COLS_SERV)


def aba_boletim(wb, aps, arquivo):
    ws = wb.create_sheet('Boletim')
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    for i, (_, w) in enumerate(COLS_SERV, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    r = 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    t = ws.cell(r, 1, 'BOLETIM DE APONTAMENTOS')
    t.font = Font(name=FONTE, size=15, bold=True, color=BRANCO)
    t.alignment = Alignment(vertical='center', indent=1)
    pinta(ws, r, NC, TINTA)
    ws.row_dimensions[r].height = 28
    r += 1

    area = sum(a['area'] for a in aps)
    ext = sum(a['ext'] for a in aps)
    datas = [a['ident'].get('Executado em') for a in aps if a['ident'].get('Executado em')]
    periodo = f'{data_br(min(datas))} a {data_br(max(datas))}' if datas else '—'
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
    s = ws.cell(r, 1, f'{len(aps)} apontamento(s) · {br(area)} m² · {br(ext, 1)} m · '
                      f'período {periodo} · origem: {arquivo} · '
                      f'emitido em {datetime.now().strftime("%d/%m/%Y %H:%M")}')
    s.font = Font(name=FONTE, size=8.5, color='FF3F4A55')
    s.alignment = Alignment(vertical='center', indent=1)
    pinta(ws, r, NC, FAIXA)
    ws.row_dimensions[r].height = 15
    r += 2

    for ap in aps:
        i = ap['ident']
        tipo = texto(ap['extras'].get('Tipo de Pintura Manual')
                     or ap['extras'].get('Tipo de Pintura Mecânica')
                     or ap['extras'].get('Item de Serviço'))
        faixa_titulo(
            ws, r, ap['serial'],
            f"{texto(i.get('Natureza'))}{' · ' + tipo if tipo else ''}"
            f"   |   {data_br(i.get('Executado em'))}   |   {texto(i.get('Status'))}",
            NC)
        r += 1

        fp, lp = i.get('Fechamento de pista'), i.get('Liberação de pista')
        interdicao = '—'
        if isinstance(fp, datetime) and isinstance(lp, datetime):
            interdicao = f'{round((lp - fp).total_seconds() / 60)} min'

        pares = [
            ('Cliente', texto(i.get('Cliente'))),
            ('Contrato', texto(i.get('Contrato'))),
            ('Local de obra', texto(i.get('Local de Obra'))),
            ('Encarregado', texto(i.get('Encarregado'))),
            ('Executado em', texto(i.get('Executado em'))),
            ('Encontrado em', texto(i.get('Encontrado em'))),
            ('Fechamento de pista', hora_br(fp)),
            ('Liberação de pista', hora_br(lp)),
            ('Tempo de interdição', interdicao),
            ('Engenheiro responsável', texto(i.get('Engenheiro responsável'))),
            ('Criado por', f"{texto(i.get('Criado por'))} · {texto(i.get('Criado em'))}"),
            ('Equipe (Kartado)', texto(i.get('Equipe'))),
        ]
        sub_faixa(ws, r, 'Identificação', NC)
        r += 1
        for k in range(0, len(pares), 2):
            ws.row_dimensions[r].height = 14
            # rótulo em 3 colunas: "ENGENHEIRO RESPONSÁVEL" tem 22 caracteres e
            # o Excel corta texto alinhado à direita sem avisar.
            par(ws, r, 1, 3, 4, 6, *pares[k])
            if k + 1 < len(pares):
                par(ws, r, 7, 9, 10, NC, *pares[k + 1])
            r += 1
        r += 1

        # ---------------- serviços ----------------
        if ap['servicos']:
            sub_faixa(ws, r, f"Serviços executados ({len(ap['servicos'])})", NC)
            r += 1
            for n, (titulo, _) in enumerate(COLS_SERV, 1):
                c = ws.cell(r, n, titulo)
                c.font = Font(name=FONTE, size=7.5, bold=True, color=CINZA)
                c.alignment = Alignment(vertical='center', horizontal='center', wrap_text=True)
                c.fill = PatternFill('solid', fgColor=CLARO)
                c.border = borda('top', 'bottom', 'left', 'right')
            ws.row_dimensions[r].height = 22
            r += 1

            ini = r
            for s in ap['servicos']:
                vals = [s['n'], s.get('kmIni'), s.get('kmFim'), s.get('cor'), s.get('local'),
                        s.get('tipoFaixa'), s.get('faixas'), s.get('ext'), s.get('larg'),
                        None, s.get('obs')]
                for n, v in enumerate(vals, 1):
                    c = ws.cell(r, n, v if v is not None else '—')
                    c.font = Font(name=FONTE, size=9, color=TINTA)
                    c.border = borda('top', 'bottom', 'left', 'right')
                    if n in (1, 2, 3, 7):
                        c.alignment = Alignment(horizontal='center')
                    elif n in (8, 9):
                        c.alignment = Alignment(horizontal='right', indent=1)
                        c.number_format = '#,##0.00'
                    else:
                        c.alignment = Alignment(horizontal='left', indent=1, wrap_text=(n == 11))

                # Área como FÓRMULA: extensão × largura × faixas. É a conta que o
                # Kartado faz, e escrevê-la aqui torna a divergência visível.
                c = ws.cell(r, 10, f'=ROUND(H{r}*I{r}*G{r},2)')
                c.number_format = '#,##0.00'
                c.alignment = Alignment(horizontal='right', indent=1)
                c.border = borda('top', 'bottom', 'left', 'right')
                decl = s.get('area')
                calc = (s.get('ext') or 0) * (s.get('larg') or 0) * (s.get('faixas') or 0)
                if decl is not None and abs(calc - decl) > 0.05:
                    c.font = Font(name=FONTE, size=9, bold=True, color=ALERTA)
                    c.comment = Comment(f'O Kartado declarou {br(decl)} m² neste serviço.', 'Incovia')
                else:
                    c.font = Font(name=FONTE, size=9, bold=True, color=TINTA)
                ws.row_dimensions[r].height = 14
                r += 1

            for n in range(1, NC + 1):
                c = ws.cell(r, n)
                c.fill = PatternFill('solid', fgColor=CLARO)
                c.border = borda('top', 'bottom', 'left', 'right')
            c = ws.cell(r, 7, 'TOTAL')
            c.font = Font(name=FONTE, size=8, bold=True, color=CINZA)
            c.alignment = Alignment(horizontal='right', indent=1)
            for col, fmt in ((8, '#,##0.0'), (10, '#,##0.00')):
                L = get_column_letter(col)
                c = ws.cell(r, col, f'=SUM({L}{ini}:{L}{r - 1})')
                c.font = Font(name=FONTE, size=9.5, bold=True, color=TINTA)
                c.number_format = fmt
                c.alignment = Alignment(horizontal='right', indent=1)
                c.border = borda('top', 'bottom', 'left', 'right')
            ws.row_dimensions[r].height = 16
            r += 2

        # ---------------- leituras (ensaios) ----------------
        if ap['leituras']:
            sub_faixa(ws, r, f"Leituras ({len(ap['leituras'])})", NC)
            r += 1
            for n, (idx, _) in enumerate(ap['leituras'], 1):
                c = ws.cell(r, n, f'm{idx}')
                c.font = Font(name=FONTE, size=8, bold=True, color=CINZA)
                c.alignment = Alignment(horizontal='center')
                c.fill = PatternFill('solid', fgColor=CLARO)
                c.border = borda('top', 'bottom', 'left', 'right')
            col_med = len(ap['leituras']) + 1
            c = ws.cell(r, col_med, 'MÉDIA')
            c.font = Font(name=FONTE, size=8, bold=True, color=TINTA)
            c.alignment = Alignment(horizontal='center')
            c.fill = PatternFill('solid', fgColor=FAIXA)
            c.border = borda('top', 'bottom', 'left', 'right')
            r += 1
            for n, (_, valor) in enumerate(ap['leituras'], 1):
                c = ws.cell(r, n, valor)
                c.font = Font(name=FONTE, size=10, color=TINTA)
                c.alignment = Alignment(horizontal='center')
                c.border = borda('top', 'bottom', 'left', 'right')
            fim = get_column_letter(len(ap['leituras']))
            c = ws.cell(r, col_med, f'=ROUND(AVERAGE(A{r}:{fim}{r}),1)')
            c.font = Font(name=FONTE, size=10, bold=True, color=TINTA)
            c.number_format = '0.0'
            c.alignment = Alignment(horizontal='center')
            c.fill = PatternFill('solid', fgColor=CLARO)
            c.border = borda('top', 'bottom', 'left', 'right')
            r += 2

        # ---------------- observações ----------------
        obs = texto(ap['extras'].get('Observações'))
        if obs:
            sub_faixa(ws, r, 'Observações do apontamento', NC)
            r += 1
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
            c = ws.cell(r, 1, obs)
            c.font = Font(name=FONTE, size=9.5, color=TINTA)
            c.alignment = Alignment(vertical='top', wrap_text=True, indent=1)
            ws.row_dimensions[r].height = 26
            r += 1

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=NC)
        c = ws.cell(r, 1, f"{ap['fotos']} foto(s) anexada(s)")
        c.font = Font(name=FONTE, size=8, italic=True, color=CINZA)
        c.alignment = Alignment(vertical='center', indent=1)
        r += 3

    return ws


# ----------------------------------------------------------------------
# Aba Dados — uma linha por serviço
# ----------------------------------------------------------------------
COLS_DADOS = [
    ('Serial', 22), ('Data', 11), ('Hora', 8), ('Natureza', 20), ('Tipo', 13),
    ('Status', 11), ('Cliente', 11), ('Contrato', 11), ('Local de Obra', 22),
    ('Encarregado', 15), ('Engenheiro', 15), ('Criado por', 17),
    ('Serviço nº', 9), ('Km inicial', 10), ('Km final', 10), ('Cor', 11),
    ('Local de aplicação', 19), ('Tipo de faixa', 15), ('Faixas', 8),
    ('Extensão (m)', 12), ('Largura (m)', 11), ('Área (m²)', 11),
    ('Obs. do serviço', 26),
]


def aba_dados(wb, aps):
    ws = wb.create_sheet('Dados')
    ws.sheet_view.showGridLines = False

    for n, (t, w) in enumerate(COLS_DADOS, 1):
        c = ws.cell(1, n, t)
        c.font = Font(name=FONTE, size=8.5, bold=True, color=BRANCO)
        c.fill = PatternFill('solid', fgColor=TINTA)
        c.alignment = Alignment(vertical='center', indent=1, wrap_text=True)
        ws.column_dimensions[get_column_letter(n)].width = w
    ws.row_dimensions[1].height = 26
    ws.freeze_panes = 'A2'

    r = 2
    for ap in aps:
        i = ap['ident']
        tipo = texto(ap['extras'].get('Tipo de Pintura Manual')
                     or ap['extras'].get('Tipo de Pintura Mecânica')
                     or ap['extras'].get('Item de Serviço'))
        base = [ap['serial'], data_br(i.get('Executado em')), hora_br(i.get('Executado em')),
                texto(i.get('Natureza')), tipo, texto(i.get('Status')),
                texto(i.get('Cliente')), texto(i.get('Contrato')),
                texto(i.get('Local de Obra')), texto(i.get('Encarregado')),
                texto(i.get('Engenheiro responsável')), texto(i.get('Criado por'))]

        # apontamento sem serviços (ensaio) ainda vira uma linha, para não sumir
        for s in (ap['servicos'] or [{}]):
            vals = base + [
                s.get('n', ''), s.get('kmIni'), s.get('kmFim'), s.get('cor'),
                s.get('local'), s.get('tipoFaixa'), s.get('faixas'),
                s.get('ext'), s.get('larg'), s.get('area'), s.get('obs'),
            ]
            for n, v in enumerate(vals, 1):
                c = ws.cell(r, n, v if v is not None else '')
                c.font = Font(name=FONTE, size=9.5, color=TINTA)
                c.border = borda('bottom')
                if isinstance(v, (int, float)) and not isinstance(v, bool):
                    c.number_format = '#,##0.00' if isinstance(v, float) else '#,##0'
                    c.alignment = Alignment(vertical='center', horizontal='right', indent=1)
                else:
                    c.alignment = Alignment(vertical='center', indent=1)
            ws.row_dimensions[r].height = 15
            r += 1

    ws.auto_filter.ref = f'A1:{get_column_letter(len(COLS_DADOS))}{max(1, r - 1)}'
    return ws


def main(entrada, saida):
    aps = ler(entrada)
    if not aps:
        print('Nenhum apontamento encontrado.')
        return
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    aba_boletim(wb, aps, entrada.split('/')[-1])
    aba_dados(wb, aps)
    wb.save(saida)
    n_serv = sum(len(a['servicos']) for a in aps)
    print(f"{len(aps)} apontamento(s) · {n_serv} serviço(s) · "
          f"{br(sum(a['area'] for a in aps))} m² · gravado em {saida}")


if __name__ == '__main__':
    main(sys.argv[1], sys.argv[2])
